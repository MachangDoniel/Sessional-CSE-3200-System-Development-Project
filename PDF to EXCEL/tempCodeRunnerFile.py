import pdfplumber
from openpyxl import load_workbook

# Open the PDF file
with pdfplumber.open('table.pdf') as pdf:
    # Extract data from the PDF
    data = {}
    for page in pdf.pages:
        for line in page.extract_text().split('\n'):
            row = line.split()  # Assuming the data is space-separated, adjust as needed
            if len(row) >= 2:
                serial_no = row[0]
                name = row[1]
                data[serial_no] = name

# Load the Excel file
excel_file = 'data from table.xlsx'
wb = load_workbook(excel_file)
sheet = wb.active  # Select the active sheet, you can choose a specific sheet if needed

# Iterate through the Excel file and insert names
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
    serial_no = row[0].value
    if serial_no in data:
        name = data[serial_no]
        row[1].value = name

# Save the modified Excel file
wb.save(excel_file)
